import os
import cv2
import numpy as np
from curve_fitting.config import *
from curve_fitting.curve_fit import *
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def findDistance(rgb_dir, segmentation_dir, output_dir, create_excel=False,create_images=True):
    
    if create_excel == True:
        wb = Workbook()
        ws = wb.active
        ws.title = "Road Edge Measurements"
        headers = ["Image Name", "Left Distance (Linear)", "Right Distance (Linear)", "Forward Distance"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
    
    rgb_files = [f for f in os.listdir(rgb_dir) if f.endswith('.jpg')]
    count = 0
    # Iterate through the files
    for rgb_file in tqdm(rgb_files, desc="Processing images"):
        # Construct full file paths
        rgb_path = os.path.join(rgb_dir, rgb_file)
        
        segmentation_file = rgb_file.replace('.jpg', '_segmentation.png') 
        segmentation_path1 = os.path.join(segmentation_dir, segmentation_file)
        
        # Check if corresponding segmentation file exists
        if not os.path.exists(segmentation_path1):
            print(f"Segmentation file not found for {rgb_file} in 1. Skipping.")
            break

        original_image = cv2.imread(rgb_path)
        segmented_image1 = cv2.imread(segmentation_path1, cv2.IMREAD_UNCHANGED)
        height = original_image.shape[0]
        crop_height = int(height * 0.80)
        original_image_cropped = original_image[:crop_height, :]
        original_image_resized = cv2.resize(original_image_cropped, (128, 128), interpolation=cv2.INTER_LINEAR)
        binary_map1 = segmented_image1.astype(np.uint8)
        scale_x = 128 / original_image_cropped.shape[1]
        scale_y = 128 / original_image_cropped.shape[0]
        original_cy = original_matrix[1,2]
        cy_cropped = (original_cy - 0) / crop_height * height  # Assuming crop starts from top (0)
        camera_matrix = np.array([[original_matrix[0,0] * scale_x, 0., original_matrix[0,2] * scale_x],
                            [0., original_matrix[1,1] * scale_y, original_matrix[1,2] * scale_y * 0.80],  # Adjust y-center for cropping
                            [0., 0., 1.]])    
        # Detect road edges and calculate distances
        ransac_min, ransac_max, edges = detect_road_edges(binary_map1)
        
        # Create visualization (using existing code)
        

        # Define the fixed y-coordinate (you can adjust this value)
        fixed_y = int(binary_map1.shape[0] * 0.90)  # 90% down the image
        camera_center_x = original_image_resized.shape[1] / 2
        camera_center_y = original_image_resized.shape[0] / 2

        # Calculate x-coordinates for left and right edges at the fixed y
        left_x = ransac_min.predict([[fixed_y]])[0]
        right_x = ransac_max.predict([[fixed_y]])[0]

        # Calculate distances for left and right edges
        left_distance, right_distance, forward_distance, left_euclidean, right_euclidean = estimate_distances_to_edges(
            left_x, right_x, camera_center_x, known_road_width, camera_matrix, camera_height, camera_pitch, fixed_y)
        if(create_images==True):
            plt.figure(figsize=(12, 8))
            plt.imshow(cv2.cvtColor(original_image_resized, cv2.COLOR_BGR2RGB))
            # Visualization (similar to your original code, with updated annotations)
            plt.axhline(y=fixed_y, color='yellow', linestyle='--', linewidth=2)
            plt.scatter(left_x, fixed_y, color='red', s=100)
            plt.scatter(right_x, fixed_y, color='green', s=100)
            plt.scatter(camera_center_x, camera_center_y, color='blue', s=100, marker='x')
            plt.plot([camera_center_x, left_x], [camera_center_y, fixed_y], color='red', linestyle='--')
            plt.plot([camera_center_x, right_x], [camera_center_y, fixed_y], color='green', linestyle='--')

            y_range = np.array(range(original_image_resized.shape[0]))
            plt.plot(ransac_min.predict(y_range.reshape(-1, 1)), y_range, color='red', linewidth=2)
            plt.plot(ransac_max.predict(y_range.reshape(-1, 1)), y_range, color='green', linewidth=2)

            plt.title('Road Edges and Distances from Car Centerline')
            plt.axis('on')

            plt.annotate(f'Left: {left_distance:.2f}m (L)', 
                        (left_x, fixed_y), xytext=(10, -20), textcoords='offset points')
            plt.annotate(f'Right: {right_distance:.2f}m (L)', 
                        (right_x, fixed_y), xytext=(10, 20), textcoords='offset points')
            plt.plot([left_x, right_x], [fixed_y, fixed_y], color='yellow', linewidth=2)

            plt.annotate('Measurement Line', (left_x, fixed_y), xytext=(10, -30), 
                        textcoords='offset points', color='yellow')
            plt.legend(['Fixed Y-line', 'Left Edge Point', 'Right Edge Point', 
                        'Left Distance', 'Right Distance', 'Left Edge', 'Right Edge'])

            output_path = os.path.join(output_dir, f"result_{rgb_file}")
            plt.savefig(output_path)
            plt.close()
        
        if create_excel == 1:
            row = ws.max_row + 1
            ws.cell(row=row, column=1, value=(rgb_file.split('_jpg', 1)[0].split('output',1)[1]))
            ws.cell(row=row, column=2, value=left_distance)
            ws.cell(row=row, column=3, value=right_distance)
            ws.cell(row=row, column=4, value=forward_distance)
        
        count += 1
    
    if create_excel == True:
        excel_output_path = os.path.join(output_dir, "road_edge_measurements.xlsx")
        wb.save(excel_output_path)
        print(f"Excel file saved at: {excel_output_path}")
    
    print("Processing complete.")


def overlay(rgb_dir,segmentation_dir,output_dir):
    rgb_files = ([f for f in os.listdir(rgb_dir) if f.endswith('.jpg')])
    count = 0
    # Iterate through the files
    for rgb_file in tqdm(rgb_files, desc="Processing images"):
        # Construct full file paths
        rgb_path = os.path.join(rgb_dir, rgb_file)
        
        segmentation_file = rgb_file.replace('.jpg', '_segmentation.png') 
        segmentation_path1 = os.path.join(segmentation_dir, segmentation_file)
        
        # Check if corresponding segmentation file exists
        if not os.path.exists(segmentation_path1):
            print(f"Segmentation file not found for {rgb_file} in 1. Skipping.")
            break


        original_image = cv2.imread(rgb_path)
        segmented_image1 = cv2.imread(segmentation_path1, cv2.IMREAD_UNCHANGED)
        # print(segmented_image1.shape)
        height = original_image.shape[0]
        crop_height = int(height * 0.80)
        original_image_cropped = original_image[:crop_height, :]
        original_image_resized = cv2.resize(original_image_cropped, (128, 128), interpolation=cv2.INTER_LINEAR)
        binary_map1 = segmented_image1.astype(np.uint8)
        # Create and save segmentation overlay
        plt.figure(figsize=(12, 8))
        plt.imshow(cv2.cvtColor(original_image_resized, cv2.COLOR_BGR2RGB))
        binary_map_255 = (binary_map1 > 0).astype(np.uint8) * 255
        colored_mask = np.zeros_like(original_image_resized)
        colored_mask[:,:,1] = binary_map_255
        overlay = cv2.addWeighted(original_image_resized, 0.7, colored_mask, 0.3, 0)
        plt.imshow(cv2.cvtColor(overlay, cv2.COLOR_BGR2RGB))
        plt.title('Original Image with Segmentation Mask Overlay')
        overlay_output_path = os.path.join(output_dir, f"overlay_{rgb_file}.png")
        plt.savefig(overlay_output_path)
        plt.close()
        count +=1
    print("Processing complete.")

def Excel(rgb_dir,segmentation_dir,output_dir):
    wb = Workbook()
    ws = wb.active
    ws.title = "Road Edge Measurements"
    headers = ["Image Name", "Left Distance (Linear)", "Right Distance (Linear)"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    rgb_files = ([f for f in os.listdir(rgb_dir) if f.endswith('.jpg')])
    count = 0
    # Iterate through the files
    for rgb_file in tqdm(rgb_files, desc="Processing images"):
        # Construct full file paths
        rgb_path = os.path.join(rgb_dir, rgb_file)
        
        segmentation_file = rgb_file.replace('.jpg', '_segmentation.png') 
        segmentation_path1 = os.path.join(segmentation_dir, segmentation_file)
        
        # Check if corresponding segmentation file exists
        if not os.path.exists(segmentation_path1):
            print(f"Segmentation file not found for {rgb_file} in 1. Skipping.")
            break


        original_image = cv2.imread(rgb_path)
        segmented_image1 = cv2.imread(segmentation_path1, cv2.IMREAD_UNCHANGED)
        # print(segmented_image1.shape)
        height = original_image.shape[0]
        crop_height = int(height * 0.80)
        original_image_cropped = original_image[:crop_height, :]
        original_image_resized = cv2.resize(original_image_cropped, (128, 128), interpolation=cv2.INTER_LINEAR)
        binary_map1 = segmented_image1.astype(np.uint8)
        scale_x = 128 / original_image_cropped.shape[1]
        scale_y = 128 / original_image_cropped.shape[0]
        original_cy = original_matrix[1,2]
        crop_height = int(height * 0.75)
        cy_cropped = (original_cy - 0) / crop_height * height  # Assuming crop starts from top (0)
        camera_matrix = np.array([[original_matrix[0,0] * scale_x, 0., original_matrix[0,2] * scale_x],
                            [0., original_matrix[1,1] * scale_y, original_matrix[1,2] * scale_y * 0.75],  # Adjust y-center for cropping
                            [0., 0., 1.]])    
        # Detect road edges and calculate distances
        ransac_min, ransac_max, edges = detect_road_edges(binary_map1)
        
        # Create visualization (using existing code)
        camera_center_x = original_image_resized.shape[1] / 2
        camera_center_y = original_image_resized.shape[0] / 2

        # Define the fixed y-coordinate (you can adjust this value)
        fixed_y = int(binary_map1.shape[0] * 0.90)  # 80% down the image

        # Calculate x-coordinates for left and right edges at the fixed y
        left_x = ransac_min.predict([[fixed_y]])[0]
        right_x = ransac_max.predict([[fixed_y]])[0]

        # Calculate distances for left and right edges
        left_distance, right_distance, forward_distance, left_euclidean, right_euclidean = estimate_distances_to_edges(
            left_x, right_x, camera_center_x, known_road_width, camera_matrix, camera_height, camera_pitch, fixed_y)
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=(rgb_file.split('_jpg', 1)[0].split('output',1)[1]))
        ws.cell(row=row, column=2, value=left_distance)
        ws.cell(row=row, column=3, value=right_distance)
        ws.cell(row=row, column=4, value=forward_distance)
        # ws.cell(row=row, column=5, value=left_euclidean)
        # ws.cell(row=row, column=6, value=right_euclidean)


        count +=1
    excel_output_path = os.path.join(output_dir, "road_edge_measurements.xlsx")
    wb.save(excel_output_path)
    print(f"Excel file saved at: {excel_output_path}")
    print("Processing complete.")

    






